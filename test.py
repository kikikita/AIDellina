import openpyxl as ox
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows

import openpyxl as ox
import pandas as pd

def update_spreadsheet(path: str, data, startcol: int = 1,
                       startrow: int = 1, sheet_name: str = 'Sheet1'):
    """
    Функция для обновления таблицы в Excel, сохраняя существующие диаграммы и стили.

    :param path: путь к файлу
    :param data: данные для вставки (может быть значением или pandas DataFrame)
    :param startcol: начальная колонка (1-based index)
    :param startrow: начальная строка (1-based index)
    :param sheet_name: имя листа
    """
    wb = ox.load_workbook(path)
    ws = wb[sheet_name]

    # Ensure charts and images stay in place
    if ws._charts:
        charts = ws._charts.copy()
    else:
        charts = []

    if ws._images:
        images = ws._images.copy()
    else:
        images = []

    # Insert data
    if isinstance(data, pd.DataFrame):
        for r_idx, row in enumerate(data.itertuples(index=False, name=None), startrow):
            for c_idx, value in enumerate(row, startcol):
                ws.cell(row=r_idx, column=c_idx, value=value)
    else:
        ws.cell(row=startrow, column=startcol, value=data)

    # Restore charts and images
    ws._charts = charts
    ws._images = images

    # Save the workbook
    wb.save(path)
