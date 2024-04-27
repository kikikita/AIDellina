import pandas as pd
import numpy as np
import pypyodbc as odbc
from darts import TimeSeries
from darts.models import Prophet,  AutoARIMA, ExponentialSmoothing
from darts.dataprocessing.transformers import Scaler
from darts.metrics import mape, r2_score
import random
from scipy.optimize import fsolve
import warnings
from datetime import datetime
from dateutil.relativedelta import relativedelta
import shutil
import openpyxl as ox
import logging
warnings.filterwarnings('ignore')
logging.disable(logging.CRITICAL)


DRIVER_NAME = 'SQL SERVER'
SERVER_NAME = 'm1-DWH1'
DATABASE_NAME = 'DataRaw'
DAYS = pd.read_excel('samples/days.xlsx')

connection_string = f"""
    DRIVER={{{DRIVER_NAME}}};
    SERVER={SERVER_NAME};
    DATABASE={DATABASE_NAME};
    Trust_Connection=yes;
"""
conn = odbc.connect(connection_string)

month_dict = {'Янв': '01', 'Фев': '02', 'Мар': '03', 'Апр': '04', 'Май': '05',
              'Июн': '06', 'Июл': '07', 'Авг': '08', 'Сен': '09', 'Окт': '10',
              'Ноя': '11', 'Дек': '12'}


def get_data(osp: str, date_start, forecast: None = None):
    """
    Функция для получения данных из базы данных и последующей их обработки.

    Параметры:
    :osp (str): строка с названиями ОСП, разделенными запятой и пробелом.
    :date_start: дата, с которой брать данные для прогноза
    :forecast: использование прогноза СПП

    Возвращает:
    df_ (DataFrame): основной DataFrame с данными за последние 60 месяцев.
    df_kg (DataFrame): DataFrame с данными по показателю 'кг нетто'.
    df_nv (DataFrame): DataFrame с данными по показателю 'накладные'.
    parts (dict): словарь с долями различных услуг.
    """
    osp = ', '.join(f"'{x.strip()}'" for x in osp.split(', '))

    base_query = f"""
    SELECT *
    FROM (
        SELECT
            [Год],
            [Месяц],
            [Показатель],
            [Услуга],
            [Итоговый]
        FROM {{}}
        WHERE [ОСПОтправитель] IN ({osp})
        AND [Показатель] IN ('кг нетто','накладные')
        AND [Год] >= 2014
    ) AS sourceTable
    PIVOT
    (
        SUM([Итоговый])
        FOR [Услуга] IN ([МТ], [Автодоставка забор], [Край забор],
        [Экспресс-доставка], [Груз МТ входящий], [Автодоставка отвоз],
        [Край отвоз], [Экспресс-доставка входящий груз])
    ) AS PivotTable
    """

    SQL_QUERY = base_query.format(
        "[DataRaw].[Mart].[Агрегированный факт грузооборота ЛТЛ]")
    df = pd.read_sql(SQL_QUERY, conn)

    if forecast == 1 or forecast == '1':
        print('Выгружаем прогноз из СПП...')

        forecast_SQL_QUERY = base_query.format(
            "[DataRaw].[Mart].[Прогноз грузооборота ЛТЛ]")
        df_forecast = pd.read_sql(forecast_SQL_QUERY, conn)
        df = pd.concat([df, df_forecast])

    df['месяц'] = df['месяц'].map(month_dict)
    df['ds'] = pd.to_datetime(df['год'].astype(str) +
                              df['месяц'] + '01', format='%Y%m%d')
    df = df.sort_values(by='ds').set_index('ds')\
        .reset_index().drop(['год', 'месяц'], axis=1)
    df = pd.merge(df, DAYS, left_on='ds', right_on='date', how='left')\
        .fillna(0).drop(['date'], axis=1)

    for col in df.columns[1:]:
        if col not in ('показатель', 'days'):
            df[col] = df[col] / df['days']
    df = df[df['ds'] >= '2014-01-01']

    df_kg = df[df['показатель'] == 'кг нетто'].drop(['показатель'], axis=1)\
        .rename(columns={'мт': 'mt_1', 'груз мт входящий': 'mt_2',
                         'автодоставка забор': 'ad_ish_kg',
                         'автодоставка отвоз': 'ad_vh_kg',
                         'край забор': 'krai_ish_kg',
                         'край отвоз': 'krai_vh_kg',
                         'экспресс-доставка': 'exp_ish_kg',
                         'экспресс-доставка входящий груз': 'exp_vh_kg'})
    df_nv = df[df['показатель'] == 'накладные'].drop(['показатель'], axis=1)\
        .rename(columns={'мт': 'nv_1', 'груз мт входящий': 'nv_2',
                         'автодоставка забор': 'ad_ish_nv',
                         'автодоставка отвоз': 'ad_vh_nv',
                         'край забор': 'krai_ish_nv',
                         'край отвоз': 'krai_vh_nv',
                         'экспресс-доставка': 'exp_ish_nv',
                         'экспресс-доставка входящий груз': 'exp_vh_nv'})

    parts = {}
    if 'ad_ish_kg' in df_kg.columns and 'ad_vh_kg' in df_kg.columns:
        parts['ad_ish_kg'] = np.mean(
            df_kg['ad_ish_kg'].tail(6)/df_kg['mt_1'].tail(6))
        parts['ad_vh_kg'] = np.mean(
            df_kg['ad_vh_kg'].tail(6)/df_kg['mt_2'].tail(6))
        parts['ad_ish_nv'] = np.mean(
            df_nv['ad_ish_nv'].tail(6)/df_nv['nv_1'].tail(6))
        parts['ad_vh_nv'] = np.mean(
            df_nv['ad_vh_nv'].tail(6)/df_nv['nv_2'].tail(6))

    if 'krai_ish_kg' in df_kg.columns:
        parts['krai_ish_kg'] = np.mean(
            df_kg['krai_ish_kg'].tail(6)/df_kg['mt_1'].tail(6))
        parts['krai_ish_nv'] = np.mean(
            df_nv['krai_ish_nv'].tail(6)/df_nv['nv_1'].tail(6))

    if 'krai_vh_kg' in df_kg.columns:
        parts['krai_vh_kg'] = np.mean(
            df_kg['krai_vh_kg'].tail(6)/df_kg['mt_2'].tail(6))
        parts['krai_vh_nv'] = np.mean(
            df_nv['krai_vh_nv'].tail(6)/df_nv['nv_2'].tail(6))

    if 'exp_ish_kg' in df_kg.columns:
        parts['exp_ish_kg'] = np.mean(
            df_kg['exp_ish_kg'].tail(6)/df_kg['mt_1'].tail(6))
        parts['exp_ish_nv'] = np.mean(
            df_nv['exp_ish_nv'].tail(6)/df_nv['nv_1'].tail(6))

    if 'exp_vh_kg' in df_kg.columns:
        parts['exp_vh_kg'] = np.mean(
            df_kg['exp_vh_kg'].tail(6)/df_kg['mt_2'].tail(6))
        parts['exp_vh_nv'] = np.mean(
            df_nv['exp_vh_nv'].tail(6)/df_nv['nv_2'].tail(6))

    for key in parts.keys():
        if parts[key] < 0.005:
            parts[key] = 0

    df_ = pd.concat([df_kg[['ds', 'mt_1', 'mt_2']].reset_index(drop=True),
                     df_nv[['nv_1', 'nv_2']].reset_index(drop=True)], axis=1)

    df_ = df_[df_['ds'] >= date_start].fillna(0)

    df_.drop(df_.loc[df['ds'] == 0].index, inplace=True)

    return df_, df_kg, df_nv, parts


def mt_coef_finder(cut_forecast: pd.DataFrame, weight: float):
    """
    Функция для поиска коэффициентов MT.

    Параметры:
    cut_forecast (pd.DataFrame): DataFrame с прогнозными значениями.
    weight (float): весовой коэффициент.

    Возвращает:
    mt_1_cf, mt_2_cf: коэффициенты mt_1 и mt_2.
    """
    def func(x):
        return [(x[0] * np.sum(cut_forecast['mt_1'])) /
                (x[0] * np.sum(cut_forecast['mt_1']) +
                 np.sum(cut_forecast['mt_2'])) - weight]
    root = fsolve(func, 1)
    min_cf = 1 - (1 - root) / 2
    max_cf = 1 + (1 - root) / 2
    mt_1_cf, mt_2_cf = min_cf, max_cf
    return mt_1_cf, mt_2_cf


def result_mt_changer(cut_df: pd.DataFrame, result: pd.DataFrame):
    """
    Функция для изменения результатов MT.

    Параметры:
    cut_df (pd.DataFrame): DataFrame с обрезанными значениями.
    result (pd.DataFrame): DataFrame с результатами.

    Возвращает:
    result: измененный DataFrame с результатами.
    """
    cut_df['summ'] = cut_df['mt_1'] + cut_df['mt_2']
    weight = np.sum(cut_df['mt_1']) / np.sum(cut_df['summ'])
    year_list = result.index.year.unique()
    for i in year_list:
        mt_1_cf, mt_2_cf = mt_coef_finder(result.loc[f'{i}'], weight)
        result.loc[f'{i}', 'mt_1'] *= mt_1_cf[0]
        result.loc[f'{i}', 'mt_2'] *= mt_2_cf[0]
    return result


def tn_coef_finder(cut_df, periods):
    """
    Функция для поиска коэффициентов TN.

    Параметры:
    cut_df: DataFrame с обрезанными значениями.
    periods: количество периодов.

    Возвращает:
    tn_coefs: словарь с коэффициентами TN.
    """
    tn_coefs = {}
    for i in range(1, ((len(cut_df.columns)+1)//2)):
        cut_df[f'mt_nv_{i}'] = cut_df[f'mt_{i}']/cut_df[f'nv_{i}']
        # tn_coefs[f'tn_{i}'] = [f'mt_{i}', np.mean(cut_df[f'mt_tn_{i}'])]
        mean_cf = np.mean(cut_df[f'mt_nv_{i}'])
        tn_coefs[f'nv_{i}'] = [f'mt_{i}',
                               [random.uniform(mean_cf-2, mean_cf+2)
                                for _ in range(periods)]]
    return tn_coefs


def stationary_r2_score(actual: np.ndarray, predicted: np.ndarray) -> float:
    # Flatten the predicted array if it's a nested list
    if np.ndim(predicted) == 2:
        predicted = predicted.flatten()

    # Compute the first difference of the time series
    actual_diff = np.diff(actual, n=1)
    predicted_diff = np.diff(predicted, n=1)

    # Compute the R2 score on the differenced data
    return r2_score(
        TimeSeries.from_values(actual_diff),
        TimeSeries.from_values(predicted_diff[:len(actual_diff)]))


def forecast(df: pd.DataFrame, periods: int):
    """
    Функция для прогнозирования.

    Параметры:
    df (pd.DataFrame): исходный DataFrame.
    periods (int): количество периодов для прогноза.

    Возвращает:
    result: DataFrame с прогнозными значениями.
    """
    # Read the data
    cut_df = df.tail(6).copy()

    # Prepare for metrics calculation
    last_6_months = df.iloc[-6:]
    model_dict = {}
    mape_dict = {}
    r2_dict = {}
    st_r2_dict = {}

    result = []
    for i in range(1, len(df.columns)):
        col_name = df.columns[i]

        temp_df = df[['ds', col_name]].copy()
        temp_df.columns = ['ds', 'y']

        scaler = Scaler()
        series = scaler.fit_transform(
            TimeSeries.from_dataframe(temp_df, 'ds', 'y'))
        train_series = scaler.transform(
            TimeSeries.from_dataframe(temp_df.iloc[:-6], 'ds', 'y'))

        models = [ExponentialSmoothing(), AutoARIMA(), Prophet()]
        best_model = None
        best_mape = float('inf')

        # Iterate over all combinations of parameters
        for model in models:
            model.fit(train_series)
            future = model.predict(6)

            # Rescale the predicted series to the original scale
            future = scaler.inverse_transform(future)

            predicted = future.values()[-6:]
            actual = last_6_months[col_name].values
            predicted_ts = TimeSeries.from_values(predicted)
            actual_ts = TimeSeries.from_values(actual)
            current_mape = mape(actual_ts, predicted_ts)
            current_r2 = r2_score(actual_ts, predicted_ts)
            current_st_r2 = stationary_r2_score(actual, predicted)

            if current_mape < best_mape:
                best_mape = current_mape
                best_r2 = current_r2
                best_st_r2 = current_st_r2
                best_model = model

        best_model.fit(series)
        future = best_model.predict(int(periods))
        future = scaler.inverse_transform(future)

        temp_df = future.pd_dataframe().reset_index()
        temp_df.columns = ['ds', col_name]

        if i == 1:
            result = temp_df
        else:
            result = result.join(temp_df.set_index('ds'), on='ds')

        model_dict[col_name] = type(best_model).__name__
        mape_dict[col_name] = best_mape
        r2_dict[col_name] = best_r2
        st_r2_dict[col_name] = best_st_r2

    result = result.tail(periods).set_index('ds')
    # result = result_mt_changer(cut_df, result)
    tn_coefs = tn_coef_finder(cut_df, periods)
    for k, v in tn_coefs.items():
        result[k] = result[v[0]] / v[1]

    df_r2 = pd.DataFrame({'R2': r2_dict})
    df_mape = pd.DataFrame({'MAPE': mape_dict})
    df_model = pd.DataFrame({'Model name': model_dict})
    metrics = pd.concat([df_r2, df_mape, df_model], axis=1)

    return result, metrics


# Вычисляет разницу между текущей датой и датой до которой нужен прогноз,
# влючая текущий и последний месяцы
def months_difference(last_date, date_until):
    """
    Функция для определения разности в месяцах между двумя датами.

    :param last_date: последняя дата
    :param date_until: дата до которой идет отсчет
    :return: разница в месяцах между датами
    """
    difference = relativedelta(last_date, date_until)
    return np.abs(difference.years * 12) + np.abs(difference.months)


def copy_sample(new_name: str, periods: int):
    """
    Функция для копирования образца файла с учетом определенного условия.

    :param new_name: новое имя файла
    :param periods: количество периодов
    :return: путь к новому файлу
    """
    source_file = 'samples/sample.xlsx'
    if periods > 75:
        source_file = 'samples/sample_10.xlsx'
        new_name = new_name + '_10_лет'
    new_file = f'data/Расчет ОСП {new_name}.xlsx'
    shutil.copy(source_file, new_file)
    return new_file


def find_start_row(path, date_to_find):
    """
    Функция для нахождения начальной строки в таблице по дате.

    :param path: путь к файлу
    :param date_to_find: дата, по которой ищется строка
    :return: индекс начальной строки
    """
    df = pd.read_excel(path, sheet_name='Расчёт (кг)')

    if date_to_find in df.iloc[:, 2].values:
        row_index = df.index[df.iloc[:, 2] == date_to_find].to_list()[0] + 2
        return row_index


def update_spreadsheet(path: str, _data, startcol: int = 1,
                       startrow: int = 1, sheet_name: str = 'Sheet1',
                       withnames: None = None):
    """
    Функция для обновления таблицы в Excel.

    :param path: путь к файлу
    :param _data: данные для вставки
    :param startcol: начальная колонка
    :param startrow: начальная строка
    :param sheet_name: имя листа
    """
    wb = ox.load_workbook(path)

    if not isinstance(_data, pd.DataFrame):
        wb[sheet_name].cell(startrow, startcol).value = _data
    else:
        if withnames is not None:
            for ic, col_name in enumerate(_data.columns):
                wb[sheet_name].cell(startrow, startcol + ic).value = col_name
            startrow += 1  # Move data one row below

        for ir in range(0, len(_data)):
            for ic in range(0, len(_data.iloc[ir])):
                wb[sheet_name].cell(startrow + ir, startcol + ic)\
                    .value = _data.iloc[ir][ic]

    wb.save(path)


def make_full_predict(osp: str, date_open, date_start,
                      date_until, is_forecast):
    """
    Функция для построения полного прогноза и сохранения в excel-файле.

    :param osp: параметр наименование ОСП
    :param date_open: дата открытия
    :param date_until: дата до которой необходим прогноз
    :param is_forecast: использовать прогноз СПП или нет
    :return: None
    """
    try:
        print("Выгрузка данных...")
        df, df_kg, df_nv, parts = get_data(osp, date_start, is_forecast)
    except Exception:
        print('Ошибка в выгрузке данных')
        return
    # try:
    print("Построение прогноза...")
    last_date = df['ds'].max()
    periods = months_difference(last_date, date_until)
    pred, metrics = forecast(df, periods)
    # except Exception:
    #     print('Ошибка в построении прогноза')

    for key in parts.keys():
        if 'ish_kg' in key:
            pred[key] = pred['mt_1'] * parts[key]
        if 'vh_kg' in key:
            pred[key] = pred['mt_2'] * parts[key]
        if 'ish_nv' in key:
            pred[key] = pred['nv_1'] * parts[key]
        if 'vh_nv' in key:
            pred[key] = pred['nv_2'] * parts[key]

    kg = pred[[col for col in pred.columns if 'mt' in col or 'kg' in col]]
    nv = pred[[col for col in pred.columns if 'nv' in col]]

    kg = pd.merge(kg.reset_index(), DAYS,
                  left_on='ds', right_on='date', how='left')\
        .fillna(0).drop(['date'], axis=1).set_index('ds')
    nv = pd.merge(nv.reset_index(), DAYS, left_on='ds',
                  right_on='date', how='left')\
        .fillna(0).drop(['date'], axis=1).set_index('ds')

    date_to_find = df_kg.iloc[0, 0]

    df_kg.set_index('ds', inplace=True)
    df_nv.set_index('ds', inplace=True)

    df_kg = pd.concat([df_kg, kg], axis=0).fillna(0)
    df_nv = pd.concat([df_nv, nv], axis=0).fillna(0)

    for df in (df_kg, df_nv):
        for col in df.columns:
            if col != 'days':
                df[col] = df[col] * df['days']

    file_path = copy_sample(osp, periods)
    start_row = find_start_row(file_path, date_to_find)

    print("Сохранение результатов...")
    update_spreadsheet(file_path, osp, sheet_name='Форма',
                       startrow=1, startcol=2)
    update_spreadsheet(file_path, date_open, sheet_name='Форма',
                       startrow=7, startcol=2)
    # update_spreadsheet(file_path, metrics.reset_index(),
    #                    sheet_name='Форма', startrow=1, startcol=5,
    #                    withnames=True)
    update_spreadsheet(file_path, df_kg.iloc[:, :-1], sheet_name='Расчёт (кг)',
                       startrow=start_row, startcol=5)
    update_spreadsheet(file_path, df_nv.iloc[:, :-1], sheet_name='Расчёт (ТН)',
                       startrow=start_row, startcol=5)

    print('Прогноз построен и сохранен в папку /data!')
    print(metrics)


if __name__ == "__main__":
    osp = str(input('Введите наименование ОСП: '))
    # dates = input(
    #     'Введите даты прогноза через пробел в формате YYYY-MM-DD ' +
    #     '(вторая дата опциональна): ')\
    #     .split()
    date_open = input('Введите дату октрытия ОСП в формате YYYY-MM-DD: ')
    date_start_ = input(
        'Введите дату, с которой стоит брать данные, ' +
        'или нажмите "Enter": ')
    date_until_ = input(
        'Введите дату, до которой необходим прогноз, ' +
        'или нажмите "Enter": ')
    is_forecast = input('Использовать прогноз СПП? Да - 1, Нет - "Enter": ')
    # try:
    date_open = datetime.strptime(date_open, "%Y-%m-%d")

    date_start = datetime.strptime("2014-01-01", "%Y-%m-%d")
    if len(date_start_) == 10 and date_start_.startswith('2'):
        date_start = datetime.strptime(date_start_, "%Y-%m-%d")

    date_until = datetime.strptime("2029-12-01", "%Y-%m-%d")
    if len(date_until_) == 10 and date_until_.startswith('2'):
        date_until = datetime.strptime(date_until_, "%Y-%m-%d")

    make_full_predict(osp, date_open, date_start, date_until, is_forecast)
    # except ValueError:
    #     print('Неверный формат даты. Используйте формат YYYY-MM-DD')
